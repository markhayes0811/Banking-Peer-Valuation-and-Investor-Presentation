#!/usr/bin/env python3
"""
Minimal helper to refresh prices in the Excel model using Polygon.
Usage:
  export POLYGON_API_KEY="your_key"
  python src/fill_from_polygon.py --excel model/Equity_Valuation_Model_v2-Hayes.xlsx
"""
import os, argparse, datetime as dt, requests
from openpyxl import load_workbook

API = os.environ.get("POLYGON_API_KEY")

def last_close(ticker: str):
    end = dt.date.today()
    start = end - dt.timedelta(days=30)
    url = f"https://api.polygon.io/v2/aggs/ticker/{ticker}/range/1/day/{start}/{end}?adjusted=true&sort=desc&limit=1&apiKey={API}"
    r = requests.get(url, timeout=15)
    r.raise_for_status()
    j = r.json()
    if j.get("resultsCount", 0) > 0:
        return j["results"][0]["c"]
    return None

def update_excel(path: str):
    wb = load_workbook(path)
    ws = wb["Inputs"]
    row_by_ticker = {ws.cell(r,2).value: r for r in range(2, ws.max_row+1)}
    for t in ("USB","WFC","SQ"):
        px = last_close(t)
        if px and t in row_by_ticker:
            r = row_by_ticker[t]
            ws.cell(r, 4).value = float(px)  # Stock Price
    wb.save(path)
    print("Updated prices in", path)

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default="model/Equity_Valuation_Model_v2-Hayes.xlsx")
    args = ap.parse_args()
    assert API, "Set POLYGON_API_KEY in your environment."
    update_excel(args.excel)
